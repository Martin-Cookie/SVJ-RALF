"""Excel import service for parsing SVJ owner data.

Based on the proven implementation from the original SVJ project.

Expected file: SVJ_Evidence_Vlastniku_CLEAN.xlsx
Sheet: Vlastnici_SVJ
Columns (0-indexed):
  A  (0)  = Číslo jednotky (KN)          e.g. "1098/1"
  B  (1)  = Číslo prostoru (stavební)     e.g. "A 111"
  C  (2)  = Podíl na SČD                  e.g. 12212
  D  (3)  = Podlahová plocha (m²)         e.g. 185.56
  E  (4)  = Počet místností               e.g. "3+1"
  F  (5)  = Druh prostoru                 e.g. "byt"
  G  (6)  = Sekce domu                    e.g. "A"
  H  (7)  = Číslo orientační              e.g. 22
  I  (8)  = Adresa jednotky               e.g. "Štěpařská"
  J  (9)  = LV číslo                      e.g. 3504
  K  (10) = Typ vlastnictví               e.g. "ANO", "VL", "SJVL"
  L  (11) = Jméno                         first name or company name
  M  (12) = Příjmení / název              last name or company name
  N  (13) = Titul                         e.g. "Ing."
  O  (14) = Rodné číslo / IČ              e.g. "711128/9911" or "12345678"
  P  (15) = Trvalá adresa – ulice
  Q  (16) = Trvalá adresa – část obce
  R  (17) = Trvalá adresa – město
  S  (18) = Trvalá adresa – PSČ
  T  (19) = Trvalá adresa – stát
  U  (20) = Koresp. adresa – ulice
  V  (21) = Koresp. adresa – část obce
  W  (22) = Koresp. adresa – město
  X  (23) = Koresp. adresa – PSČ
  Y  (24) = Koresp. adresa – stát
  Z  (25) = Telefon GSM
  AA (26) = Telefon pevný
  AB (27) = Email (Evidence 2024)
  AC (28) = Email (Kontakty)
  AD (29) = Vlastník od
  AE (30) = Poznámka
"""
from __future__ import annotations

import re
from unicodedata import normalize, category

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.owner import Owner, OwnerUnit, Unit

# Column indices (0-based)
COL_UNIT_KN = 0
COL_BUILDING_NUM = 1
COL_PODIL_SCD = 2
COL_FLOOR_AREA = 3
COL_ROOM_COUNT = 4
COL_SPACE_TYPE = 5
COL_SECTION = 6
COL_ORIENT_NUM = 7
COL_ADDRESS = 8
COL_LV_NUMBER = 9
COL_OWNERSHIP_TYPE = 10
COL_FIRST_NAME = 11
COL_LAST_NAME = 12
COL_TITLE = 13
COL_BIRTH_OR_IC = 14
COL_PERM_STREET = 15
COL_PERM_DISTRICT = 16
COL_PERM_CITY = 17
COL_PERM_ZIP = 18
COL_PERM_COUNTRY = 19
COL_CORR_STREET = 20
COL_CORR_DISTRICT = 21
COL_CORR_CITY = 22
COL_CORR_ZIP = 23
COL_CORR_COUNTRY = 24
COL_PHONE_GSM = 25
COL_PHONE_LANDLINE = 26
COL_EMAIL_EVIDENCE = 27
COL_EMAIL_CONTACTS = 28
COL_OWNER_SINCE = 29
COL_NOTE = 30

SHEET_NAME = "Vlastnici_SVJ"


def _cell(row: tuple, idx: int) -> str | None:
    """Safely get cell value as stripped string, or None."""
    if idx >= len(row) or row[idx] is None:
        return None
    val = str(row[idx]).strip()
    return val if val else None


def _cell_int(row: tuple, idx: int) -> int | None:
    """Safely get cell value as integer."""
    raw = _cell(row, idx)
    if raw is None:
        return None
    try:
        return int(float(raw))
    except (ValueError, TypeError):
        return None


def _cell_float(row: tuple, idx: int) -> float | None:
    """Safely get cell value as float."""
    raw = _cell(row, idx)
    if raw is None:
        return None
    try:
        return float(raw)
    except (ValueError, TypeError):
        return None


def _strip_diacritics(text: str) -> str:
    """Remove diacritics from text."""
    nfkd = normalize("NFD", text)
    return "".join(c for c in nfkd if category(c) != "Mn")


def _normalize_name(text: str) -> str:
    """Normalize name for matching: lowercase, no diacritics, single spaces."""
    result = _strip_diacritics(text.lower())
    return " ".join(result.split())


def _is_birth_number(value: str) -> bool:
    """Check if value looks like Czech birth number (XXXXXX/XXXX or 10 digits)."""
    clean = value.replace(" ", "")
    if "/" in clean:
        parts = clean.split("/")
        return len(parts) == 2 and parts[0].isdigit() and len(parts[0]) == 6 and parts[1].isdigit()
    return clean.isdigit() and len(clean) == 10


def _is_company_id(value: str) -> bool:
    """Check if value looks like Czech IČ (8-digit number, no slash)."""
    clean = value.replace(" ", "")
    return clean.isdigit() and len(clean) == 8 and "/" not in value


def _detect_owner_type(birth_or_ic: str | None) -> str:
    """Detect owner type: 'legal' if IČ (8-digit), otherwise 'physical'."""
    if birth_or_ic and _is_company_id(birth_or_ic):
        return "legal"
    return "physical"


_TRAILING_FRACTION_RE = re.compile(r"\s+\d+/\d+$")


def _clean_name(value: str | None) -> str | None:
    """Strip trailing fraction patterns (e.g. '1/3') from names."""
    if not value:
        return value
    return _TRAILING_FRACTION_RE.sub("", value).strip() or value


def _normalize_ownership_type(raw: str | None) -> str | None:
    """Normalize ownership type: ANO -> SJM, keep others as-is."""
    if not raw:
        return None
    val = raw.strip()
    if val.upper() == "ANO":
        return "SJM"
    return val


def _build_name_with_titles(title: str | None, first_name: str, last_name: str | None) -> str:
    """Build display name: title + příjmení + jméno.

    Deduplicates when first_name == last_name (common for legal entities).
    """
    parts = []
    if title:
        parts.append(title)
    if last_name:
        parts.append(last_name)
    if first_name and first_name != last_name:
        parts.append(first_name)
    return " ".join(parts)


def _build_name_normalized(first_name: str, last_name: str | None) -> str:
    """Build normalized name for search: 'last_name first_name' lowercased, no diacritics."""
    parts = []
    if last_name:
        parts.append(last_name)
    if first_name:
        parts.append(first_name)
    return _normalize_name(" ".join(parts))


def _owner_group_key(first_name: str | None, last_name: str | None, birth_or_ic: str | None) -> str:
    """Generate a grouping key for identifying unique owners across rows.

    Uses birth_number/IČ if available, otherwise normalized name.
    """
    if birth_or_ic:
        clean = birth_or_ic.replace(" ", "").strip()
        if clean:
            return f"id:{clean}"
    fn = _normalize_name(first_name or "")
    ln = _normalize_name(last_name or "")
    return f"name:{ln}|{fn}"


def _describe_skip_error(row: tuple, row_idx: int) -> str:
    """Build an error message for a skipped row."""
    unit_kn = _cell(row, COL_UNIT_KN)
    first_name = _cell(row, COL_FIRST_NAME)
    last_name = _cell(row, COL_LAST_NAME)

    missing = []
    if not unit_kn:
        missing.append("číslo jednotky")
    if not first_name:
        missing.append("jméno")

    present = []
    if unit_kn:
        present.append(f"jednotka={unit_kn}")
    if first_name:
        present.append(f"jméno={first_name}")
    if last_name:
        present.append(f"příjmení={last_name}")
    title = _cell(row, COL_TITLE)
    if title:
        present.append(f"titul={title}")
    birth_ic = _cell(row, COL_BIRTH_OR_IC)
    if birth_ic:
        present.append(f"RČ/IČ={birth_ic}")

    msg = f"Řádek {row_idx}: chybí {', '.join(missing)}"
    if present:
        msg += f" (nalezeno: {', '.join(present)})"
    return msg


def _get_worksheet(file_bytes_or_path):
    """Open workbook and return the correct worksheet."""
    wb = load_workbook(file_bytes_or_path, read_only=True, data_only=True)
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active
    return wb, ws


def _parse_unit_number(unit_kn_raw: str) -> int | None:
    """Parse unit number from Excel: '1098/1' -> 1, '115' -> 115."""
    if "/" in unit_kn_raw:
        unit_kn_raw = unit_kn_raw.split("/")[-1].strip()
    try:
        return int(unit_kn_raw)
    except (ValueError, TypeError):
        return None


def _parse_row(row: tuple, row_idx: int) -> dict | None:
    """Parse a single row into a structured dict. Returns None if row should be skipped."""
    unit_kn_raw = _cell(row, COL_UNIT_KN)
    first_name = _cell(row, COL_FIRST_NAME)

    if not unit_kn_raw or not first_name:
        return None

    unit_kn = _parse_unit_number(unit_kn_raw)
    if unit_kn is None:
        return None

    return {
        "row_idx": row_idx,
        # Unit data
        "unit_kn": unit_kn,
        "unit_kn_raw": unit_kn_raw,
        "building_number": _cell(row, COL_BUILDING_NUM),
        "podil_scd": _cell_int(row, COL_PODIL_SCD),
        "floor_area": _cell_float(row, COL_FLOOR_AREA),
        "room_count": _cell(row, COL_ROOM_COUNT),
        "space_type": _cell(row, COL_SPACE_TYPE),
        "section": _cell(row, COL_SECTION),
        "orientation_number": _cell_int(row, COL_ORIENT_NUM),
        "address": _cell(row, COL_ADDRESS),
        "lv_number": _cell_int(row, COL_LV_NUMBER),
        # Owner data
        "ownership_type": _cell(row, COL_OWNERSHIP_TYPE),
        "first_name": first_name,
        "last_name": _clean_name(_cell(row, COL_LAST_NAME)),
        "title": _cell(row, COL_TITLE),
        "birth_or_ic": _cell(row, COL_BIRTH_OR_IC),
        # Addresses
        "perm_street": _cell(row, COL_PERM_STREET),
        "perm_district": _cell(row, COL_PERM_DISTRICT),
        "perm_city": _cell(row, COL_PERM_CITY),
        "perm_zip": _cell(row, COL_PERM_ZIP),
        "perm_country": _cell(row, COL_PERM_COUNTRY),
        "corr_street": _cell(row, COL_CORR_STREET),
        "corr_district": _cell(row, COL_CORR_DISTRICT),
        "corr_city": _cell(row, COL_CORR_CITY),
        "corr_zip": _cell(row, COL_CORR_ZIP),
        "corr_country": _cell(row, COL_CORR_COUNTRY),
        # Contacts
        "phone_gsm": _cell(row, COL_PHONE_GSM),
        "phone_landline": _cell(row, COL_PHONE_LANDLINE),
        "email_evidence": _cell(row, COL_EMAIL_EVIDENCE),
        "email_contacts": _cell(row, COL_EMAIL_CONTACTS),
        # Other
        "owner_since": _cell(row, COL_OWNER_SINCE),
        "note": _cell(row, COL_NOTE),
    }


def preview_owners_from_excel(file_bytes_or_path) -> dict:
    """Parse Excel and return preview data without saving to DB."""
    wb, ws = _get_worksheet(file_bytes_or_path)

    owner_keys = set()
    unit_numbers = set()
    rows_processed = 0
    errors = []
    preview_rows = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        parsed = _parse_row(row, row_idx)
        if parsed is None:
            if row and any(c is not None for c in row[:15]):
                errors.append(_describe_skip_error(row, row_idx))
            continue

        rows_processed += 1
        unit_numbers.add(parsed["unit_kn"])

        key = _owner_group_key(parsed["first_name"], parsed["last_name"], parsed["birth_or_ic"])
        owner_keys.add(key)

        owner_type = _detect_owner_type(parsed["birth_or_ic"])

        last = parsed["last_name"] or ""
        first = parsed["first_name"] or ""
        preview_rows.append({
            "row": row_idx,
            "name": _build_name_with_titles(parsed["title"], first, last),
            "sort_name": f"{last} {first}".strip().lower(),
            "owner_type": owner_type,
            "unit_number": parsed["unit_kn"],
            "unit_kn_raw": parsed["unit_kn_raw"],
            "building_number": parsed["building_number"] or "",
            "ownership_type": _normalize_ownership_type(parsed["ownership_type"]) or "",
            "podil_scd": parsed["podil_scd"] or 0,
            "section": parsed["section"] or "",
            "email": parsed["email_evidence"] or parsed["email_contacts"] or "",
            "phone": parsed["phone_gsm"] or "",
        })

    wb.close()

    return {
        "rows_processed": rows_processed,
        "owners_count": len(owner_keys),
        "units_count": len(unit_numbers),
        "preview_rows": preview_rows,
        "errors": errors,
    }


def import_owners_from_excel(db: Session, file_bytes_or_path) -> dict:
    """Parse Excel and save owners, units, and relationships to DB."""
    wb, ws = _get_worksheet(file_bytes_or_path)

    # First pass: collect all rows grouped by owner key
    owner_groups: dict[str, list[dict]] = {}
    rows_processed = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        parsed = _parse_row(row, row_idx)
        if parsed is None:
            if row and any(c is not None for c in row[:15]):
                errors.append(_describe_skip_error(row, row_idx))
            continue

        rows_processed += 1
        key = _owner_group_key(parsed["first_name"], parsed["last_name"], parsed["birth_or_ic"])
        owner_groups.setdefault(key, []).append(parsed)

    wb.close()

    # Second pass: create DB records
    owners_created = 0
    units_created = 0
    links_created = 0
    unit_cache: dict[int, Unit] = {}

    for key, rows in owner_groups.items():
        # Pick the row with the cleanest last_name (shortest = least noise)
        first_row = min(rows, key=lambda r: len(r["last_name"] or ""))

        # Detect owner type
        owner_type = _detect_owner_type(first_row["birth_or_ic"])

        # Parse birth number vs company ID
        birth_number = None
        company_id_val = None
        birth_or_ic = first_row["birth_or_ic"]
        if birth_or_ic:
            if _is_company_id(birth_or_ic):
                company_id_val = birth_or_ic.strip()
            elif _is_birth_number(birth_or_ic):
                birth_number = birth_or_ic.strip()
            else:
                birth_number = birth_or_ic.strip()

        # Build names
        name_with_titles = _build_name_with_titles(
            first_row["title"], first_row["first_name"], first_row["last_name"]
        )
        name_normalized = _build_name_normalized(first_row["first_name"], first_row["last_name"])

        # Pick best email/phone from all rows for this owner
        email = None
        email_secondary = None
        phone = None
        phone_landline = None
        for r in rows:
            if not email and r["email_evidence"]:
                email = r["email_evidence"]
            if not email_secondary and r["email_contacts"]:
                email_secondary = r["email_contacts"]
            if not phone and r["phone_gsm"]:
                phone = r["phone_gsm"]
            if not phone_landline and r["phone_landline"]:
                phone_landline = r["phone_landline"]

        owner = Owner(
            first_name=first_row["first_name"],
            last_name=first_row["last_name"],
            title=first_row["title"],
            name_with_titles=name_with_titles,
            name_normalized=name_normalized,
            owner_type=owner_type,
            birth_number=birth_number,
            company_id=company_id_val,
            perm_street=first_row["perm_street"],
            perm_district=first_row["perm_district"],
            perm_city=first_row["perm_city"],
            perm_zip=first_row["perm_zip"],
            perm_country=first_row["perm_country"],
            corr_street=first_row["corr_street"],
            corr_district=first_row["corr_district"],
            corr_city=first_row["corr_city"],
            corr_zip=first_row["corr_zip"],
            corr_country=first_row["corr_country"],
            phone=phone,
            phone_landline=phone_landline,
            email=email,
            email_secondary=email_secondary,
            owner_since=first_row["owner_since"],
            note=first_row["note"],
        )
        db.add(owner)
        db.flush()
        owners_created += 1

        # Create units and owner-unit links
        for row_data in rows:
            unit_kn = row_data["unit_kn"]

            if unit_kn not in unit_cache:
                existing_unit = db.query(Unit).filter_by(unit_number=unit_kn).first()
                if existing_unit:
                    unit_cache[unit_kn] = existing_unit
                else:
                    unit = Unit(
                        unit_number=unit_kn,
                        building_number=row_data["building_number"],
                        podil_scd=row_data["podil_scd"],
                        floor_area=row_data["floor_area"],
                        room_count=row_data["room_count"],
                        space_type=row_data["space_type"],
                        section=row_data["section"],
                        orientation_number=row_data["orientation_number"],
                        address=row_data["address"],
                        lv_number=row_data["lv_number"],
                    )
                    db.add(unit)
                    db.flush()
                    unit_cache[unit_kn] = unit
                    units_created += 1

            unit_obj = unit_cache[unit_kn]
            votes_val = unit_obj.podil_scd or 0

            owner_unit = OwnerUnit(
                owner_id=owner.id,
                unit_id=unit_obj.id,
                ownership_type=_normalize_ownership_type(row_data["ownership_type"]),
                share=1.0,
                votes=votes_val,
                excel_row_number=row_data["row_idx"],
            )
            db.add(owner_unit)
            links_created += 1

    # NOTE: caller is responsible for db.commit() — allows transactional control
    db.flush()

    return {
        "owners_created": owners_created,
        "units_created": units_created,
        "links_created": links_created,
        "rows_processed": rows_processed,
        "errors": errors,
    }
