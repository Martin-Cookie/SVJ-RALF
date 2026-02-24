"""Excel export service for owners."""
import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def export_owners_xlsx(owners: List) -> io.BytesIO:
    """Export a list of Owner objects to an Excel .xlsx file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vlastníci"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "ID", "Příjmení", "Jméno", "Titul před", "Titul za",
        "Typ", "RČ/IČ", "Email", "Telefon",
        "Trvalá - ulice", "Trvalá - město", "Trvalá - PSČ",
        "Korespondenční - ulice", "Korespondenční - město", "Korespondenční - PSČ",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, owner in enumerate(owners, 2):
        rc_ic = owner.birth_number if owner.owner_type == "fyzická" else owner.ico
        values = [
            owner.id,
            owner.last_name,
            owner.first_name,
            owner.title_before or "",
            owner.title_after or "",
            owner.owner_type,
            rc_ic or "",
            owner.email or "",
            owner.phone or "",
            owner.perm_street or "",
            owner.perm_city or "",
            owner.perm_zip or "",
            owner.corr_street or "",
            owner.corr_city or "",
            owner.corr_zip or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
