"""Voting management routes."""
import io
import os
import shutil
import uuid
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from sqlalchemy.orm import Session

from app.auth import get_current_user
from app.config import settings
from app.database import get_db
from app.models.voting import Voting, VotingItem, Ballot, BallotVote

_IMPORT_TEMP_DIR = os.path.join(settings.UPLOAD_DIR, "_voting_import_temp")
os.makedirs(_IMPORT_TEMP_DIR, exist_ok=True)

router = APIRouter()


@router.get("/hlasovani", response_class=HTMLResponse)
def voting_list(
    request: Request,
    status: str = "",
    db: Session = Depends(get_db),
):
    """List all votings with status filter."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    query = db.query(Voting)
    if status:
        query = query.filter(Voting.status == status)
    query = query.order_by(Voting.created_at.desc())
    votings = query.all()

    # Count by status for filter bubbles
    total = db.query(Voting).count()
    koncept_count = db.query(Voting).filter(Voting.status == "koncept").count()
    aktivni_count = db.query(Voting).filter(Voting.status == "aktivní").count()
    uzavrene_count = db.query(Voting).filter(Voting.status == "uzavřené").count()
    zrusene_count = db.query(Voting).filter(Voting.status == "zrušené").count()

    return request.app.state.templates.TemplateResponse(
        request,
        "voting/list.html",
        {
            "user": user,
            "votings": votings,
            "status": status,
            "total": total,
            "koncept_count": koncept_count,
            "aktivni_count": aktivni_count,
            "uzavrene_count": uzavrene_count,
            "zrusene_count": zrusene_count,
        },
    )


@router.get("/hlasovani/nova", response_class=HTMLResponse)
def voting_create_page(request: Request, db: Session = Depends(get_db)):
    """Show new voting creation form."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    return request.app.state.templates.TemplateResponse(
        request, "voting/create.html", {"user": user}
    )


@router.post("/hlasovani/nova/nahled-metadat")
async def voting_preview_metadata(
    request: Request,
    template: UploadFile = File(...),
):
    """AJAX endpoint: extract metadata from uploaded .docx and return JSON."""
    if not template.filename or not template.filename.endswith(".docx"):
        return JSONResponse({"error": "Nahrajte .docx soubor"}, status_code=400)

    upload_dir = os.path.join(settings.UPLOAD_DIR, "templates")
    os.makedirs(upload_dir, exist_ok=True)
    temp_path = os.path.join(upload_dir, f"_preview_{uuid.uuid4()}.docx")

    try:
        content = await template.read()
        with open(temp_path, "wb") as f:
            f.write(content)

        from app.services.word_parser import parse_voting_items, extract_voting_metadata

        meta = extract_voting_metadata(temp_path)
        items = parse_voting_items(temp_path)
        return JSONResponse({
            "meta": meta,
            "items_count": len(items),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse({"error": f"{type(e).__name__}: {e}"}, status_code=500)
    finally:
        try:
            os.unlink(temp_path)
        except Exception:
            pass


@router.post("/hlasovani/nova")
async def voting_create(
    request: Request,
    name: str = Form(...),
    quorum: float = Form(50.0),
    start_date: str = Form(""),
    end_date: str = Form(""),
    template: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    """Create a new voting, optionally with a .docx template."""
    from datetime import date

    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    v = Voting(name=name, quorum=quorum, status="koncept")

    if start_date:
        try:
            v.start_date = date.fromisoformat(start_date)
        except ValueError:
            pass
    if end_date:
        try:
            v.end_date = date.fromisoformat(end_date)
        except ValueError:
            pass

    db.add(v)
    db.flush()  # Get the ID before saving template

    # Handle template upload
    if template and template.filename and template.filename.endswith(".docx"):
        upload_dir = os.path.join(settings.UPLOAD_DIR, "templates")
        os.makedirs(upload_dir, exist_ok=True)
        template_file = os.path.join(upload_dir, f"voting-{v.id}.docx")
        content = await template.read()
        with open(template_file, "wb") as f:
            f.write(content)
        v.template_path = template_file

        # Parse items from template
        try:
            from app.services.word_parser import parse_voting_items

            items = parse_voting_items(template_file)
            for i, text in enumerate(items, 1):
                db.add(VotingItem(voting_id=v.id, number=i, text=text))
        except Exception:
            pass  # Template parsing is best-effort; items can be added manually

    db.commit()

    request.session["flash"] = {"type": "success", "message": f"Hlasování '{name}' vytvořeno."}
    return RedirectResponse(url=f"/hlasovani/{v.id}", status_code=303)


@router.get("/hlasovani/{voting_id}", response_class=HTMLResponse)
def voting_detail(
    voting_id: int, request: Request, db: Session = Depends(get_db)
):
    """Show voting detail with items and results."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    items = (
        db.query(VotingItem)
        .filter(VotingItem.voting_id == voting_id)
        .order_by(VotingItem.number)
        .all()
    )

    # Calculate results per item
    results = []
    for item in items:
        votes = (
            db.query(BallotVote)
            .filter(BallotVote.voting_item_id == item.id)
            .all()
        )
        pro = sum(1 for v in votes if v.vote == "PRO")
        proti = sum(1 for v in votes if v.vote == "PROTI")
        zdrzel = sum(1 for v in votes if v.vote == "Zdržel se")
        total_votes = pro + proti + zdrzel
        results.append({
            "item": item,
            "pro": pro,
            "proti": proti,
            "zdrzel": zdrzel,
            "total": total_votes,
            "pro_pct": round(pro / total_votes * 100, 1) if total_votes > 0 else 0,
            "proti_pct": round(proti / total_votes * 100, 1) if total_votes > 0 else 0,
            "zdrzel_pct": round(zdrzel / total_votes * 100, 1) if total_votes > 0 else 0,
        })

    ballot_count = db.query(Ballot).filter(Ballot.voting_id == voting_id).count()
    processed_count = (
        db.query(Ballot)
        .filter(Ballot.voting_id == voting_id, Ballot.status == "zpracován")
        .count()
    )

    return request.app.state.templates.TemplateResponse(
        request,
        "voting/detail.html",
        {
            "user": user,
            "voting": voting,
            "items": items,
            "results": results,
            "ballot_count": ballot_count,
            "processed_count": processed_count,
        },
    )


@router.post("/hlasovani/{voting_id}/pridat-bod")
def voting_add_item(
    voting_id: int,
    request: Request,
    text: str = Form(...),
    db: Session = Depends(get_db),
):
    """Add an item to a voting (only in koncept state)."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)
    if voting.status != "koncept":
        request.session["flash"] = {"type": "error", "message": "Body lze přidávat pouze v konceptu."}
        return RedirectResponse(url=f"/hlasovani/{voting_id}", status_code=303)

    # Get next number
    max_num = (
        db.query(VotingItem.number)
        .filter(VotingItem.voting_id == voting_id)
        .order_by(VotingItem.number.desc())
        .first()
    )
    next_num = (max_num[0] + 1) if max_num else 1

    item = VotingItem(voting_id=voting_id, number=next_num, text=text)
    db.add(item)
    db.commit()

    request.session["flash"] = {"type": "success", "message": f"Bod {next_num} přidán."}
    return RedirectResponse(url=f"/hlasovani/{voting_id}", status_code=303)


@router.post("/hlasovani/{voting_id}/smazat-bod/{item_id}")
def voting_delete_item(
    voting_id: int,
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Delete an item from a voting (only in koncept state)."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None or voting.status != "koncept":
        return RedirectResponse(url=f"/hlasovani/{voting_id}", status_code=303)

    item = db.query(VotingItem).filter(
        VotingItem.id == item_id, VotingItem.voting_id == voting_id
    ).first()
    if item:
        db.delete(item)
        db.commit()
        request.session["flash"] = {"type": "success", "message": "Bod smazán."}

    return RedirectResponse(url=f"/hlasovani/{voting_id}", status_code=303)


@router.post("/hlasovani/{voting_id}/stav")
def voting_change_status(
    voting_id: int,
    request: Request,
    status: str = Form(...),
    db: Session = Depends(get_db),
):
    """Change voting status."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    valid_transitions = {
        "koncept": ["aktivní", "zrušené"],
        "aktivní": ["uzavřené", "zrušené"],
        "uzavřené": [],
        "zrušené": [],
    }

    if status in valid_transitions.get(voting.status, []):
        voting.status = status
        db.commit()
        request.session["flash"] = {"type": "success", "message": f"Stav změněn na '{status}'."}
    else:
        request.session["flash"] = {"type": "error", "message": f"Nelze změnit stav z '{voting.status}' na '{status}'."}

    return RedirectResponse(url=f"/hlasovani/{voting_id}", status_code=303)


@router.post("/hlasovani/{voting_id}/generovat")
def voting_generate_ballots(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Generate ballot documents for all owners with units."""
    from app.models.owner import Owner, OwnerUnit

    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)
    if voting.status != "koncept":
        request.session["flash"] = {"type": "error", "message": "Lístky lze generovat pouze v konceptu."}
        return RedirectResponse(url=f"/hlasovani/{voting_id}", status_code=303)

    items = (
        db.query(VotingItem)
        .filter(VotingItem.voting_id == voting_id)
        .order_by(VotingItem.number)
        .all()
    )
    if not items:
        request.session["flash"] = {"type": "error", "message": "Hlasování nemá žádné body."}
        return RedirectResponse(url=f"/hlasovani/{voting_id}", status_code=303)

    # Find owners with active units
    active_owners = (
        db.query(Owner)
        .join(OwnerUnit, Owner.id == OwnerUnit.owner_id)
        .filter(OwnerUnit.valid_to.is_(None))
        .distinct()
        .all()
    )

    if not active_owners:
        request.session["flash"] = {"type": "error", "message": "Žádní vlastníci s aktivními jednotkami."}
        return RedirectResponse(url=f"/hlasovani/{voting_id}", status_code=303)

    item_texts = [item.text for item in items]
    generated = 0

    from app.models.owner import Unit
    from app.services.pdf_generator import generate_ballot_pdf

    for owner in active_owners:
        # Get owner's active units
        owner_units = (
            db.query(OwnerUnit)
            .filter(OwnerUnit.owner_id == owner.id, OwnerUnit.valid_to.is_(None))
            .all()
        )
        unit_ids = [ou.unit_id for ou in owner_units]
        units = db.query(Unit).filter(Unit.id.in_(unit_ids)).all()
        unit_numbers = ", ".join(str(u.unit_number) for u in units)

        # Generate document
        output_dir = os.path.join(settings.GENERATED_DIR, f"voting-{voting_id}")
        output_path = os.path.join(output_dir, f"ballot-{voting_id}-{owner.id}.docx")

        generate_ballot_pdf(
            template_path=voting.template_path or "",
            output_path=output_path,
            voting_name=voting.name,
            owner_name=owner.display_name,
            unit_numbers=unit_numbers,
            items=item_texts,
        )

        # Create ballot record for first unit
        ballot = Ballot(
            voting_id=voting_id,
            owner_id=owner.id,
            unit_id=units[0].id if units else None,
            status="vygenerován",
            pdf_path=output_path,
        )
        db.add(ballot)
        generated += 1

    voting.status = "aktivní"
    db.commit()

    request.session["flash"] = {"type": "success", "message": f"Vygenerováno {generated} lístků. Hlasování aktivováno."}
    return RedirectResponse(url=f"/hlasovani/{voting_id}", status_code=303)


@router.get("/hlasovani/{voting_id}/listky", response_class=HTMLResponse)
def voting_ballots_list(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """List ballots for a voting."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    ballots = (
        db.query(Ballot)
        .filter(Ballot.voting_id == voting_id)
        .all()
    )

    return request.app.state.templates.TemplateResponse(
        request,
        "voting/ballots.html",
        {
            "user": user,
            "voting": voting,
            "ballots": ballots,
        },
    )


@router.get("/hlasovani/{voting_id}/listek/{ballot_id}", response_class=HTMLResponse)
def ballot_detail(
    voting_id: int,
    ballot_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Show ballot detail with vote form."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    ballot = db.query(Ballot).filter(
        Ballot.id == ballot_id, Ballot.voting_id == voting_id
    ).first()
    if ballot is None:
        return HTMLResponse("Lístek nenalezen", status_code=404)

    items = (
        db.query(VotingItem)
        .filter(VotingItem.voting_id == voting_id)
        .order_by(VotingItem.number)
        .all()
    )

    # Get existing votes for this ballot
    existing_votes = {
        bv.voting_item_id: bv.vote
        for bv in db.query(BallotVote).filter(BallotVote.ballot_id == ballot_id).all()
    }

    return request.app.state.templates.TemplateResponse(
        request,
        "voting/ballot_detail.html",
        {
            "user": user,
            "voting": voting,
            "ballot": ballot,
            "items": items,
            "existing_votes": existing_votes,
        },
    )


@router.post("/hlasovani/{voting_id}/zpracovat/{ballot_id}")
async def process_single_ballot(
    voting_id: int,
    ballot_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Process a single ballot — record votes."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    ballot = db.query(Ballot).filter(
        Ballot.id == ballot_id, Ballot.voting_id == voting_id
    ).first()
    if ballot is None:
        return HTMLResponse("Lístek nenalezen", status_code=404)

    items = (
        db.query(VotingItem)
        .filter(VotingItem.voting_id == voting_id)
        .all()
    )

    # Delete existing votes for this ballot
    db.query(BallotVote).filter(BallotVote.ballot_id == ballot_id).delete()

    # Record new votes from form data
    form = await request.form()

    for item in items:
        vote_key = f"vote_{item.id}"
        vote_value = form.get(vote_key, "")
        if vote_value in ("PRO", "PROTI", "Zdržel se"):
            db.add(BallotVote(
                ballot_id=ballot_id,
                voting_item_id=item.id,
                vote=vote_value,
            ))

    ballot.status = "zpracován"
    db.commit()

    request.session["flash"] = {"type": "success", "message": "Lístek zpracován."}
    return RedirectResponse(url=f"/hlasovani/{voting_id}/listky", status_code=303)


@router.get("/hlasovani/{voting_id}/zpracovani", response_class=HTMLResponse)
def processing_page(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Show ballot processing interface."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    ballots = (
        db.query(Ballot)
        .filter(Ballot.voting_id == voting_id)
        .all()
    )

    items = (
        db.query(VotingItem)
        .filter(VotingItem.voting_id == voting_id)
        .order_by(VotingItem.number)
        .all()
    )

    return request.app.state.templates.TemplateResponse(
        request,
        "voting/processing.html",
        {
            "user": user,
            "voting": voting,
            "ballots": ballots,
            "items": items,
        },
    )


@router.post("/hlasovani/{voting_id}/zpracovat-hromadne")
async def process_bulk_ballots(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Bulk process multiple ballots with same votes."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    items = (
        db.query(VotingItem)
        .filter(VotingItem.voting_id == voting_id)
        .all()
    )

    # Parse form data
    form = await request.form()
    ballot_ids = form.getlist("ballot_ids")

    processed = 0
    for bid_str in ballot_ids:
        bid = int(bid_str)
        ballot = db.query(Ballot).filter(
            Ballot.id == bid, Ballot.voting_id == voting_id
        ).first()
        if ballot is None:
            continue

        # Delete existing votes
        db.query(BallotVote).filter(BallotVote.ballot_id == bid).delete()

        # Record votes
        for item in items:
            vote_key = f"vote_{item.id}"
            vote_value = form.get(vote_key, "")
            if vote_value in ("PRO", "PROTI", "Zdržel se"):
                db.add(BallotVote(
                    ballot_id=bid,
                    voting_item_id=item.id,
                    vote=vote_value,
                ))

        ballot.status = "zpracován"
        processed += 1

    db.commit()

    request.session["flash"] = {"type": "success", "message": f"Zpracováno {processed} lístků."}
    return RedirectResponse(url=f"/hlasovani/{voting_id}/zpracovani", status_code=303)


@router.get("/hlasovani/{voting_id}/neodevzdane", response_class=HTMLResponse)
def unsubmitted_ballots(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """List unprocessed ballots."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    ballots = (
        db.query(Ballot)
        .filter(
            Ballot.voting_id == voting_id,
            Ballot.status.notin_(["zpracován"]),
        )
        .all()
    )

    return request.app.state.templates.TemplateResponse(
        request,
        "voting/unsubmitted.html",
        {
            "user": user,
            "voting": voting,
            "ballots": ballots,
        },
    )


def _require_editor_voting(request: Request, db: Session):
    """Check that current user has editor or admin role."""
    user = get_current_user(request, db)
    if user is None:
        return None, RedirectResponse(url="/login", status_code=303)
    if user.role not in ("admin", "editor"):
        request.session["flash"] = {"type": "error", "message": "Nemáte oprávnění pro import."}
        return None, RedirectResponse(url="/hlasovani", status_code=303)
    return user, None


_MAX_UPLOAD_SIZE = 10 * 1024 * 1024  # 10 MB
_ALLOWED_EXTENSIONS = {".xlsx", ".xls"}


@router.get("/hlasovani/{voting_id}/import", response_class=HTMLResponse)
def voting_import_page(
    voting_id: int, request: Request, db: Session = Depends(get_db)
):
    """Show voting result import page."""
    user, redirect = _require_editor_voting(request, db)
    if redirect:
        return redirect

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    items = db.query(VotingItem).filter(VotingItem.voting_id == voting_id).order_by(VotingItem.number).all()

    return request.app.state.templates.TemplateResponse(
        request,
        "voting/import.html",
        {"user": user, "voting": voting, "items": items, "preview": None},
    )


@router.post("/hlasovani/{voting_id}/import")
async def voting_import_upload(
    voting_id: int,
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Upload Excel with voting results, show preview."""
    user, redirect = _require_editor_voting(request, db)
    if redirect:
        return redirect

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    items = db.query(VotingItem).filter(VotingItem.voting_id == voting_id).order_by(VotingItem.number).all()

    # Validate file extension
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in _ALLOWED_EXTENSIONS:
        request.session["flash"] = {"type": "error", "message": "Neplatný formát souboru. Povolené: .xlsx, .xls"}
        return RedirectResponse(url=f"/hlasovani/{voting_id}/import", status_code=303)

    # Read and validate file size
    content = await file.read()
    if len(content) > _MAX_UPLOAD_SIZE:
        request.session["flash"] = {"type": "error", "message": "Soubor je příliš velký (max 10 MB)."}
        return RedirectResponse(url=f"/hlasovani/{voting_id}/import", status_code=303)

    # Save uploaded file to temp directory
    token = str(uuid.uuid4())
    temp_path = os.path.join(_IMPORT_TEMP_DIR, f"{token}.xlsx")
    # Path traversal validation (token is UUID4, but defense in depth)
    real_path = os.path.realpath(temp_path)
    real_dir = os.path.realpath(_IMPORT_TEMP_DIR)
    if not real_path.startswith(real_dir + os.sep):
        request.session["flash"] = {"type": "error", "message": "Neplatná cesta souboru."}
        return RedirectResponse(url=f"/hlasovani/{voting_id}/import", status_code=303)

    with open(temp_path, "wb") as f:
        f.write(content)

    # Parse Excel for preview
    import openpyxl
    try:
        wb = openpyxl.load_workbook(temp_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
    except Exception as e:
        request.session["flash"] = {"type": "error", "message": f"Chyba čtení souboru: {e}"}
        return RedirectResponse(url=f"/hlasovani/{voting_id}/import", status_code=303)

    headers = [str(h) if h else "" for h in rows[0]] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []

    # Store token for confirm
    request.session["voting_import_token"] = token

    preview = {
        "headers": headers,
        "rows": data_rows[:20],  # Preview first 20
        "total": len(data_rows),
        "matched": 0,
        "unmatched": 0,
    }

    # Try to match rows to owners/units
    from app.models.owner import Owner, Unit
    matched = 0
    for row in data_rows:
        if len(row) >= 2:
            owner_val = str(row[0]) if row[0] else ""
            unit_val = str(row[1]) if row[1] else ""
            # Try to find matching ballot
            if unit_val:
                try:
                    unit_num = int(unit_val)
                    unit = db.query(Unit).filter(Unit.unit_number == unit_num).first()
                    if unit:
                        matched += 1
                except (ValueError, TypeError):
                    pass

    preview["matched"] = matched
    preview["unmatched"] = len(data_rows) - matched

    return request.app.state.templates.TemplateResponse(
        request,
        "voting/import.html",
        {"user": user, "voting": voting, "items": items, "preview": preview},
    )


@router.post("/hlasovani/{voting_id}/import/potvrdit")
def voting_import_confirm(
    voting_id: int, request: Request, db: Session = Depends(get_db)
):
    """Confirm voting result import from Excel."""
    user, redirect = _require_editor_voting(request, db)
    if redirect:
        return redirect

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting is None:
        return HTMLResponse("Hlasování nenalezeno", status_code=404)

    token = request.session.pop("voting_import_token", "")
    if not token:
        request.session["flash"] = {"type": "error", "message": "Žádná data k importu."}
        return RedirectResponse(url=f"/hlasovani/{voting_id}/import", status_code=303)

    temp_path = os.path.join(_IMPORT_TEMP_DIR, f"{token}.xlsx")
    if not os.path.exists(temp_path):
        request.session["flash"] = {"type": "error", "message": "Soubor importu nenalezen."}
        return RedirectResponse(url=f"/hlasovani/{voting_id}/import", status_code=303)

    items = db.query(VotingItem).filter(VotingItem.voting_id == voting_id).order_by(VotingItem.number).all()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(temp_path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # Skip header
        wb.close()

        from app.models.owner import Owner, Unit, OwnerUnit

        imported = 0
        vote_mapping = {"pro": "PRO", "1": "PRO", "ano": "PRO", "yes": "PRO",
                        "proti": "PROTI", "0": "PROTI", "ne": "PROTI", "no": "PROTI",
                        "zdržel": "Zdržel se", "zdržel se": "Zdržel se", "abstain": "Zdržel se"}

        for row in rows:
            if not row or len(row) < 3:
                continue
            owner_val = str(row[0]).strip() if row[0] else ""
            unit_val = str(row[1]).strip() if row[1] else ""

            # Find unit
            unit = None
            if unit_val:
                try:
                    unit = db.query(Unit).filter(Unit.unit_number == int(unit_val)).first()
                except (ValueError, TypeError):
                    pass

            if not unit:
                continue

            # Find ballot for this unit
            ballot = db.query(Ballot).filter(
                Ballot.voting_id == voting_id,
                Ballot.unit_id == unit.id,
            ).first()

            if not ballot:
                continue

            # Process votes (columns 2+ map to voting items)
            for i, item in enumerate(items):
                col_idx = i + 2
                if col_idx < len(row) and row[col_idx] is not None:
                    vote_raw = str(row[col_idx]).strip().lower()
                    vote_val = vote_mapping.get(vote_raw, "")
                    if vote_val:
                        # Check if vote already exists
                        existing = db.query(BallotVote).filter(
                            BallotVote.ballot_id == ballot.id,
                            BallotVote.voting_item_id == item.id,
                        ).first()
                        if existing:
                            existing.vote = vote_val
                        else:
                            bv = BallotVote(ballot_id=ballot.id, voting_item_id=item.id, vote=vote_val)
                            db.add(bv)

            ballot.status = "zpracován"
            imported += 1

        db.commit()
        request.session["flash"] = {"type": "success", "message": f"Importováno {imported} hlasovacích lístků."}

    except Exception as e:
        db.rollback()
        request.session["flash"] = {"type": "error", "message": f"Chyba importu: {e}"}
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return RedirectResponse(url=f"/hlasovani/{voting_id}", status_code=303)


@router.post("/hlasovani/{voting_id}/smazat")
def voting_delete(
    voting_id: int,
    request: Request,
    db: Session = Depends(get_db),
):
    """Delete a voting with cascade."""
    user = get_current_user(request, db)
    if user is None:
        return RedirectResponse(url="/login", status_code=303)

    voting = db.query(Voting).filter(Voting.id == voting_id).first()
    if voting:
        db.delete(voting)
        db.commit()
        request.session["flash"] = {"type": "success", "message": "Hlasování smazáno."}

    return RedirectResponse(url="/hlasovani", status_code=303)
